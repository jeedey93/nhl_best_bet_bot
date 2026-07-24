"""
Upload poolers guide projections from Excel to Supabase.

Usage:
    source .venv/bin/activate
    python scripts/upload_poolers_guide.py

What this does:
  - For players already in the DB: updates proj_gp, proj_g, proj_a, proj_pts,
    aav, upside, notes, rank, age, team, pos only.
  - For new players: inserts them (last_gp/g/a/pts, tier, risk, etc. left null).
  - Never deletes rows or touches last_*, tier, risk, pp_pct, bust_alert.

Requires the upside column to exist:
    ALTER TABLE poolers_players ADD COLUMN IF NOT EXISTS upside int;
"""

import openpyxl
import requests
import sys
import os

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "docs", "poolers-guide",
    "Guide du pooler - 100% Pool - 2026-2027.xlsx"
)

SUPABASE_URL = "https://fifurqlitkywtmhgtzeu.supabase.co"
SUPABASE_KEY = (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9"
    ".eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZpZnVycWxpdGt5d3RtaGd0emV1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY3MDIyMjQsImV4cCI6MjA5MjI3ODIyNH0"
    ".KPVPj1qwbSJJMyLR_-AhDcRs0vi2sUU6qbFQ-kH53C0"
)
HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
}

# French position abbreviations → English
POS_MAP = {
    "AD": "RW", "AG": "LW",
    "AD/AG": "RW", "AG/AD": "LW",
    "AD/C": "RW", "AG/C": "LW",
    "C/AD": "C",  "C/AG": "C",
    "LW/RW": "LW",
    "C": "C", "LW": "LW", "RW": "RW", "D": "D", "G": "G",
}


def normalize_pos(raw, fallback="F"):
    if not raw:
        return fallback
    key = str(raw).strip().upper()
    return POS_MAP.get(key, key.split("/")[0][:2])


def compute_risk(bandaid, risque):
    b = str(bandaid).strip().upper() if bandaid else ''
    r = str(risque).strip().upper() if risque else ''
    is_injured = b == 'OUI' or b == 'C'
    is_risky   = r == 'OUI'
    is_medium  = r == 'MOYEN'
    if is_risky:
        return 'High'
    if is_injured and is_medium:
        return 'High'
    if is_injured:
        return 'Medium'
    if is_medium:
        return 'Medium'
    return 'Low'


def parse_salary(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return round(val / 1_000_000, 3)
    s = str(val).replace("$", "").replace("M", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_players():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    players = []

    # ── FORWARDS ──
    ws = wb["Attaquants"]
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if r[0] is None:
            continue
        try:
            int(float(r[0]))
        except (TypeError, ValueError):
            continue
        name = str(r[1]).strip() if r[1] else None
        if not name:
            continue
        players.append({
            "name":     name,
            "age":      int(float(r[2])) if r[2] else None,
            "pos":      normalize_pos(r[3], "F"),
            "team":     str(r[4]).strip() if r[4] else None,
            "proj_gp":  int(float(r[5])) if r[5] else None,
            "proj_g":   int(float(r[6])) if r[6] else None,
            "proj_a":   int(float(r[7])) if r[7] else None,
            "proj_pts": (int(float(r[6])) if r[6] else 0) + (int(float(r[7])) if r[7] else 0),
            "aav":      parse_salary(r[10]),
            "risk":     compute_risk(r[11], r[12]),
            "upside":   int(float(r[13])) if r[13] else None,
            "notes":    str(r[14]).strip() if r[14] else None,
        })

    # ── DEFENCEMEN ──
    ws = wb["Défenseurs"]
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if r[0] is None:
            continue
        try:
            int(float(r[0]))
        except (TypeError, ValueError):
            continue
        name = str(r[1]).strip() if r[1] else None
        if not name:
            continue
        players.append({
            "name":     name,
            "age":      int(float(r[2])) if r[2] else None,
            "pos":      "D",
            "team":     str(r[4]).strip() if r[4] else None,
            "proj_gp":  int(float(r[5])) if r[5] else None,
            "proj_g":   int(float(r[6])) if r[6] else None,
            "proj_a":   int(float(r[7])) if r[7] else None,
            "proj_pts": (int(float(r[6])) if r[6] else 0) + (int(float(r[7])) if r[7] else 0),
            "aav":      parse_salary(r[10]),
            "risk":     compute_risk(r[11], r[12]),
            "upside":   int(float(r[13])) if r[13] else None,
            "notes":    str(r[14]).strip() if r[14] else None,
        })

    # ── GOALIES ──
    ws = wb["Gardiens"]
    for r in list(ws.iter_rows(values_only=True))[2:]:
        if r[0] is None:
            continue
        name = str(r[0]).strip()
        if not name:
            continue
        upside = int(float(r[15])) if r[15] else None
        players.append({
            "name":     name,
            "age":      int(float(r[1])) if r[1] else None,
            "pos":      "G",
            "team":     str(r[3]).strip() if r[3] else "FA",
            "proj_gp":  int(float(r[4])) if r[4] else None,
            "proj_g":   None,
            "proj_a":   None,
            "proj_pts": upside,
            "aav":      parse_salary(r[12]),
            "risk":     compute_risk(r[13], r[14]),
            "upside":   upside,
            "notes":    str(r[16]).strip() if r[16] else None,
        })

    # Global rank by proj_pts descending (nulls last), then sheet order as tiebreak.
    # This means rank reflects overall value across all positions.
    players_with_pts = [(p["proj_pts"] or 0, i, p) for i, p in enumerate(players)]
    players_with_pts.sort(key=lambda x: (-x[0], x[1]))
    for rank, (_, _, p) in enumerate(players_with_pts, start=1):
        p["rank"] = rank

    return players


def check_upside_column():
    res = requests.get(
        f"{SUPABASE_URL}/rest/v1/poolers_players?select=upside&limit=1",
        headers=HEADERS,
    )
    if res.status_code == 400 and "upside" in res.text:
        print("\n⚠️  The 'upside' column is missing. Run this in the Supabase SQL editor:\n")
        print("    ALTER TABLE poolers_players ADD COLUMN IF NOT EXISTS upside int;\n")
        sys.exit(1)


def fetch_existing():
    """Return {name_lower: id} for all rows currently in the DB."""
    res = requests.get(
        f"{SUPABASE_URL}/rest/v1/poolers_players?select=id,name",
        headers=HEADERS,
    )
    if res.status_code != 200:
        print(f"Failed to fetch existing players: {res.status_code} {res.text}")
        sys.exit(1)
    return {row["name"].lower(): row["id"] for row in res.json()}


def upload(players):
    existing = fetch_existing()
    print(f"  {len(existing)} players already in DB.")

    excel_names = {p["name"].lower() for p in players}

    to_update = []
    to_insert = []
    to_delete = [(name, pid) for name, pid in existing.items() if name not in excel_names]
    for p in players:
        if p["name"].lower() in existing:
            to_update.append((existing[p["name"].lower()], p))
        else:
            to_insert.append(p)

    print(f"  {len(to_update)} to update, {len(to_insert)} to insert, {len(to_delete)} to delete.")

    # Projection-only fields — never touch last_*, tier, risk, pp_pct, bust_alert
    PROJ_FIELDS = ["rank", "name", "age", "pos", "team",
                   "proj_gp", "proj_g", "proj_a", "proj_pts",
                   "aav", "risk", "upside", "notes"]

    # PATCH existing players one by one (Supabase REST doesn't bulk-patch by id list)
    updated = 0
    for pid, p in to_update:
        patch = {k: p[k] for k in PROJ_FIELDS}
        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/poolers_players?id=eq.{pid}",
            headers={**HEADERS, "Prefer": "return=minimal"},
            json=patch,
        )
        if res.status_code not in (200, 204):
            print(f"  PATCH failed for {p['name']}: {res.status_code} {res.text[:200]}")
            sys.exit(1)
        updated += 1
        if updated % 50 == 0:
            print(f"  Updated {updated}/{len(to_update)}…")
    if to_update:
        print(f"  Updated {updated}/{len(to_update)}.")

    # INSERT new players in batches of 200
    if to_insert:
        BATCH = 200
        inserted = 0
        for i in range(0, len(to_insert), BATCH):
            batch = to_insert[i : i + BATCH]
            res = requests.post(
                f"{SUPABASE_URL}/rest/v1/poolers_players",
                headers={**HEADERS, "Prefer": "return=minimal"},
                json=batch,
            )
            if res.status_code not in (200, 201):
                print(f"  INSERT failed at row {i}: {res.status_code} {res.text[:300]}")
                sys.exit(1)
            inserted += len(batch)
            print(f"  Inserted {inserted}/{len(to_insert)}…")

    # DELETE players no longer in the Excel
    if to_delete:
        BATCH = 100
        deleted = 0
        delete_ids = [pid for _, pid in to_delete]
        for i in range(0, len(delete_ids), BATCH):
            batch = delete_ids[i:i + BATCH]
            id_list = ",".join(str(x) for x in batch)
            res = requests.delete(
                f"{SUPABASE_URL}/rest/v1/poolers_players?id=in.({id_list})",
                headers=HEADERS,
            )
            if res.status_code not in (200, 204):
                print(f"  DELETE failed: {res.status_code} {res.text[:200]}")
                sys.exit(1)
            deleted += len(batch)
            print(f"  Deleted {deleted}/{len(delete_ids)}…")
        print(f"  Deleted {deleted}/{len(delete_ids)}.")

    print(f"\nDone! {len(to_update)} updated, {len(to_insert)} inserted, {len(to_delete)} deleted. Last-year stats and tier/risk untouched.")


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    print("Parsing Excel…")
    players = parse_players()
    f = sum(1 for p in players if p["pos"] not in ("D", "G"))
    d = sum(1 for p in players if p["pos"] == "D")
    g = sum(1 for p in players if p["pos"] == "G")
    print(f"  {len(players)} players parsed  ({f} F, {d} D, {g} G)")
    print()
    print("Sample:")
    for p in players[:3]:
        print(f"  #{p['rank']:3d} {p['name']:<25} {p['pos']:<3} {p['team']:<4}  "
              f"GP:{p['proj_gp']} G:{p['proj_g']} A:{p['proj_a']} PTS:{p['proj_pts']}  "
              f"AAV:${p['aav']}M  Upside:{p['upside']}")

    print()
    confirm = input("Upload to Supabase? Projections updated, players not in Excel deleted, last-year stats/tier/risk untouched. [y/N] ")
    if confirm.lower() != "y":
        print("Aborted.")
        sys.exit(0)

    check_upside_column()
    upload(players)


if __name__ == "__main__":
    main()
